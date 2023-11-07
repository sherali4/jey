from openpyxl import Workbook
import sqlite3
database = sqlite3.connect('bot.sqlite')
cursor = database.cursor()

def add_excel():
    wb = Workbook()
    wb.create_sheet("tovar", 0)
    # insert at first position
    # grab the active worksheet
    ws = wb.active
    # ws.title = "New Title"
    # print(wb.sheetnames)
    # Data can be assigned directly to cells
    # ws['A1'] = 42
    #
    # # Rows can also be appended
    # ws.append([1, 2, 3])

    # Python types will automatically be converted
    # import datetime
    # ws['A2'] = datetime.datetime.now()

    # Save the file

    # d = ws.cell(row=4, column=2, value=10)
    # for x in range(1, 101):
    #     for y in range(1, 101):
    #         ws.cell(row=x, column=y, value=25)
    birinchi_satr = ['id', 'name', 'ulchov', 'narx', 'tarifi', 'photo', 'date', 'datetime', 'user_id', 'status', 'turi']
    # print(birinchi_satr)
    # d = ws.cell(row=1, column=1, value='id')
    row = 1
    column = 0
    for x in birinchi_satr:
        column = column + 1
        ws.cell(row=row, column=column, value=x)


    cursor.execute("SELECT * FROM 'tovar'")
    tovar = cursor.fetchall()


    for y in tovar:
        ranges = range(0, len(y))
        row = row +1
        for w in ranges:
            ws.cell(row=row, column=w + 1, value=y[w])



    wb.save("tovar.xlsx")
    wb.close()

def add_users_excel():
    wb = Workbook()
    wb.create_sheet("users", 0)
    ws = wb.active
    birinchi_satr = ['id', 'name', 'numb', 'lang', 'long', 'lat', 'location_name', 'respublika', 'viloyat', 'tuman', 'hudud']
    row = 1
    column = 0
    for x in birinchi_satr:
        column = column + 1
        ws.cell(row=row, column=column, value=x)
    cursor.execute("SELECT * FROM 'users'")
    tovar = cursor.fetchall()

    for y in tovar:
        ranges = range(0, len(y))
        row = row + 1
        for w in ranges:
            ws.cell(row=row, column=w + 1, value=y[w])
    wb.save("users.xlsx")
    wb.close()


def add_zakaz_excel():
    wb = Workbook()
    wb.create_sheet("zakaz", 0)
    ws = wb.active
    birinchi_satr = ['id', 'id_client', 'id_tovar', 'datetime', 'sotib_olish', 'delete_user']
    row = 1
    column = 0
    for x in birinchi_satr:
        column = column + 1
        ws.cell(row=row, column=column, value=x)
    cursor.execute("SELECT * FROM 'zakaz'")
    tovar = cursor.fetchall()

    for y in tovar:
        ranges = range(0, len(y))
        row = row +1
        for w in ranges:
            ws.cell(row=row, column=w + 1, value=y[w])
    wb.save("zakaz.xlsx")
    wb.close()